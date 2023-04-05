from ast import Or
from concurrent.futures import thread
from fileinput import filename
from msilib import datasizemask
import tkinter
from tracemalloc import start
from typing import Counter, List
import win32com.client as client  # gives access to outlook
import os
import openpyxl
import time
from datetime import datetime, timedelta
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import csv
from tkinter import filedialog
import pyperclip
from threading import Thread, Timer
import requests
from tkinter import scrolledtext




# Hello, this app is built by quiroalv@
# This code is simple but works. Bear with me, I am just starting on coding and this has been done mostly in Office Hours, while I was doing my daily duties as scheduler
# This script looks long but it has only 3 parts
            #1. Parametres tab
            #2. Trial DS surge
            #3. Main tab



#tkinter app with tabs
root = Tk()
root.title("Flex Surge UK")

notebooksurge = ttk.Notebook(root)
notebooksurge.grid(row=0,column=0)


#adding frames for each tab
framemain = Frame(notebooksurge)
framemain.grid(row=0,column=0)
frammeparametres = Frame(notebooksurge)
frammeparametres.grid(row=0,column=0)
frammetrialDS = Frame(notebooksurge)
frammetrialDS.grid(row=0,column=0)
notebooksurge.add(framemain, text = "Main")
notebooksurge.add(frammeparametres, text = "Parametres")
notebooksurge.add(frammetrialDS, text = "Trial DS")


user_login = os.getenv("username")



#######################################################################################
#############################    PARAMETRES TAB #######################################
#######################################################################################

#getting data in "Parametres" tab
os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\EU_Flex\PendingBlocks')
root.iconbitmap("flexiconapp.ico")


#getting parametres to use them as filters
def getlatestparametres():
    os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\EU_Flex\PendingBlocks')
    # os.chdir(os.getcwd())
    global NDfillrate
    global NDstarttimesurge
    global SDfillrate
    global SDstarttimesurge
    global NDfillratetrial
    global NDstarttimesurgetrial
    global SDfillratetrial
    global SDstarttimesurgetrial
    global C1servicetype
    global C1bulkservicetype
    global SDservicetype
    global Adhocservicetype
    global RTSservicetype
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "Manual Surge - ND/Adhoc/RTS" == row[0]:
            ManualSurgeND = row
            NDfillrate = ManualSurgeND[1]
            NDstarttimesurge = ManualSurgeND[2]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "Manual Surge - SD" == row[0]:
            ManualSurgeSD = row
            SDfillrate = ManualSurgeSD[1]
            SDstarttimesurge = ManualSurgeSD[2]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "Trial DS - ND/Adhoc/RTS" == row[0]:
            SurgeNDtrial = row
            NDfillratetrial = SurgeNDtrial[1]
            NDstarttimesurgetrial = SurgeNDtrial[2]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "Trial DS - SD" == row[0]:
            SurgeSDtrial = row
            SDfillratetrial = SurgeSDtrial[1]
            SDstarttimesurgetrial = SurgeSDtrial[2]

    #service types
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "C1 Service Type" == row[0]:
            C1servicetype = row
            C1servicetype = C1servicetype[1]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "C1 Bulk Packages Service Type" == row[0]:
            C1bulkservicetype = row
            C1bulkservicetype = C1bulkservicetype[1]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "SD Service Type" == row[0]:
            SDservicetype = row
            SDservicetype = SDservicetype[1]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "Adhoc Service Type" == row[0]:
            Adhocservicetype = row
            Adhocservicetype = Adhocservicetype[1]
    surgemetrics = csv.reader(open("Surge Parametres.csv","r"))
    for row in surgemetrics:
        if "RTS Service Type" == row[0]:
            RTSservicetype = row
            RTSservicetype = RTSservicetype[1]

# setting up parametres tabs
def parametrestabwidgets():
    global NDfillrateentry
    global SDfillrateentry
    global NDstarttimesurgeentry
    global SDstarttimesurgeentry
    global NDfillratetrialentry
    global SDfillratetrialentry
    global NDstarttimesurgetrialentry
    global SDstarttimesurgetrialentry
    global C1servicetypeentry
    global C1bulkservicetypeentry
    global SDservicetypeentry
    global Adhocservicetypeentry
    global RTSservicetypeentry

    Parametrestitle = Label(frammeparametres, text= "For Manual Surge DS", font = "Helvetica 9 bold"). grid(row=0, column=0,pady=8)
    NDfillratelabel = Label(frammeparametres, text="ND/Adhoc/RTS fill rate (%): ").grid(row=1, column=0)
    NDfillrateentry = Entry(frammeparametres, width=10)
    NDfillrateentry.grid(row= 1, column=1)
    NDfillrateentry.insert(END, NDfillrate)

    SDfillratelabel = Label(frammeparametres, text="SD fill rate (%): ").grid(row=2, column=0)
    SDfillrateentry = Entry(frammeparametres, width=10)
    SDfillrateentry.grid(row= 2, column=1)
    SDfillrateentry.insert(END, SDfillrate)

    NDstarttimesurgelabel = Label(frammeparametres, text="ND buffer to start surging (min): ").grid(row=3, column=0)
    NDstarttimesurgeentry = Entry(frammeparametres, width=10)
    NDstarttimesurgeentry.grid(row= 3, column=1)
    NDstarttimesurgeentry.insert(END, NDstarttimesurge)

    SDstarttimesurgelabel = Label(frammeparametres, text="SD buffer to start surging (min): ").grid(row=4, column=0)
    SDstarttimesurgeentry = Entry(frammeparametres, width=10)
    SDstarttimesurgeentry.grid(row= 4, column=1)
    SDstarttimesurgeentry.insert(END, SDstarttimesurge)

    ## DS in trial
    Parametrestrial = Label(frammeparametres, text= "For trial DS. Warnings will pop up", font = "Helvetica 9 bold"). grid(row=5, column=0,pady=8)
    NDfillratetriallabel = Label(frammeparametres, text="Trial ND/Adhoc/RTS fill rate (%): ").grid(row=6, column=0)
    NDfillratetrialentry = Entry(frammeparametres, width=10)
    NDfillratetrialentry.grid(row= 6, column=1)
    NDfillratetrialentry.insert(END, NDfillratetrial)
    SDfillratetriallabel = Label(frammeparametres, text="Trial SD fill rate (%): ").grid(row=7, column=0)
    SDfillratetrialentry = Entry(frammeparametres, width=10)
    SDfillratetrialentry.grid(row= 7, column=1)
    SDfillratetrialentry.insert(END, SDfillratetrial)

    NDstarttimesurgetriallabel = Label(frammeparametres, text="Trial ND buffer to start surging (min): ").grid(row=8, column=0)
    NDstarttimesurgetrialentry = Entry(frammeparametres, width=10)
    NDstarttimesurgetrialentry.grid(row= 8, column=1)
    NDstarttimesurgetrialentry.insert(END, NDstarttimesurgetrial)

    SDstarttimesurgetriallabel = Label(frammeparametres, text="Trial SD buffer to start surging (min): ").grid(row=9, column=0)
    SDstarttimesurgetrialentry = Entry(frammeparametres, width=10)
    SDstarttimesurgetrialentry.grid(row= 9, column=1)
    SDstarttimesurgetrialentry.insert(END, SDstarttimesurgetrial)

    #Service types
    Parametrestrial = Label(frammeparametres, text= "Service Types", font = "Helvetica 9 bold"). grid(row=0, column=4,pady=8)
    Spaceservicetype = Label(frammeparametres, text="       ").grid(row=1, column=3)
    C1servicetypelabel = Label(frammeparametres, text="Cycle 1 Service Type: ").grid(row=1, column=4)
    C1servicetypeentry = Entry(frammeparametres, width=23)
    C1servicetypeentry.grid(row= 1, column=5)
    C1servicetypeentry.insert(END, C1servicetype)

    C1bulkservicetypelabel = Label(frammeparametres, text="Cycle 1 Bulk Packages: ").grid(row=2, column=4)
    C1bulkservicetypeentry = Entry(frammeparametres, width=23)
    C1bulkservicetypeentry.grid(row= 2, column=5)
    C1bulkservicetypeentry.insert(END, C1bulkservicetype)

    SDservicetypelabel = Label(frammeparametres, text="SameDay Service Type: ").grid(row=3, column=4)
    SDservicetypeentry = Entry(frammeparametres, width=23)
    SDservicetypeentry.grid(row= 3, column=5)
    SDservicetypeentry.insert(END, SDservicetype)

    Adhocservicetypelabel = Label(frammeparametres, text="AD_HOC Service Type: ").grid(row=4, column=4)
    Adhocservicetypeentry = Entry(frammeparametres, width=23)
    Adhocservicetypeentry.grid(row= 4, column=5)
    Adhocservicetypeentry.insert(END, Adhocservicetype)

    RTSservicetypelabel = Label(frammeparametres, text="RTS Service Type: ").grid(row=5, column=4)
    RTSservicetypeentry = Entry(frammeparametres, width=23)
    RTSservicetypeentry.grid(row= 5, column=5)
    RTSservicetypeentry.insert(END, RTSservicetype)

    Updateparametres = Button(frammeparametres, text= "Update", width=35, command= updateparametres)
    Updateparametres.grid(row=8, column=4, columnspan = 2)

def updateparametres():
    try:
        #getting all fields
        tuptitle = ("Cycle", "Fill Rate (%)", "Surge start time buffer(mins)")

        NDfillrate = NDfillrateentry.get()
        NDstarttimesurge = NDstarttimesurgeentry.get()
        ManualSurgeND = ("Manual Surge - ND/Adhoc/RTS", NDfillrate, NDstarttimesurge)
        
        SDfillrate = SDfillrateentry.get()
        SDstarttimesurge = SDstarttimesurgeentry.get()
        ManualSurgeSD = ("Manual Surge - SD", SDfillrate, SDstarttimesurge)
        
        NDfillratetrial = NDfillratetrialentry.get()
        NDstarttimesurgetrial = NDstarttimesurgetrialentry.get()
        SurgeNDtrial = ("Trial DS - ND/Adhoc/RTS", NDfillratetrial, NDstarttimesurgetrial)

        SDfillratetrial = SDfillratetrialentry.get()
        SDstarttimesurgetrial = SDstarttimesurgetrialentry.get()
        SurgeSDtrial = ("Trial DS - SD", SDfillratetrial, SDstarttimesurgetrial)
        
        tuptitleservicetype = ("Cycle", "Service Type")
        C1servicetype = C1servicetypeentry.get()
        C1servicetypetup = ("C1 Service Type", C1servicetype)
        C1bulkservicetype = C1bulkservicetypeentry.get()
        C1bulkservicetypetup = ("C1 Bulk Packages Service Type", C1bulkservicetype)
        SDservicetype = SDservicetypeentry.get()
        SDservicetypetup = ("SD Service Type", SDservicetype)
        Adhocservicetype = Adhocservicetypeentry.get()
        Adhocservicetypetup = ("Adhoc Service Type", Adhocservicetype)
        RTSservicetype = RTSservicetypeentry.get()
        RTSservicetypetup = ("RTS Service Type", RTSservicetype)
        RTSservicetype = RTSservicetypeentry.get()
        RTSservicetypetup = ("RTS Service Type", RTSservicetype)

        f = open("Surge Parametres.csv","w",newline='')
        writer = csv.writer(f)
        writer.writerow(tuptitle)
        writer.writerow(ManualSurgeND)
        writer.writerow(ManualSurgeSD)
        writer.writerow(SurgeNDtrial)
        writer.writerow(SurgeSDtrial)
        
        writer.writerow(tuptitleservicetype)
        writer.writerow(C1servicetypetup)
        writer.writerow(C1bulkservicetypetup)
        writer.writerow(SDservicetypetup)
        writer.writerow(Adhocservicetypetup)
        writer.writerow(RTSservicetypetup)

        f.close()
        
        data = {"Content": "/md \n**Parametres Updated by " +user_login +": **\n\n**"+ str(tuptitle) + "**\n"+ str(ManualSurgeND) + "\n"+ str(ManualSurgeSD) + "\n"+ str(SurgeNDtrial) + "\n"+ str(SurgeSDtrial) + "\n\n**"+ str(tuptitleservicetype) + "**\n"+ str(C1servicetypetup) + "\n"+ str(C1bulkservicetypetup) + "\n"+ str(SDservicetypetup) + "\n"+ str(Adhocservicetypetup) + "\n"+ str(RTSservicetypetup) }
        
        sendwebhookmanagers(data)
        
        messagebox.showinfo('Success', 'Parametres have been updated')
    except:
        messagebox.showinfo('Error', 'Please ensure Surge Metrics file exists and is not opened by anyone \n File location: Business Analyses\CentralOPS\EU_Flex\PendingBlocks')











#######################################################################################
#############################    Trial DS tab #######################################
#######################################################################################

#setting up the DS that are in trial and therefore must be filtered out
def trialDSwidgets():
    global previewtrialDS
    previewtrialDS = Listbox(frammetrialDS)
    previewtrialDS.grid(row=1, column=0 ,padx=20)


    importDStrialDSbutton = Button(frammetrialDS, text= "Import file", command=importDStrial)
    importDStrialDSbutton.grid(row=0, column=0, pady=20,padx=20)

def importDStrial():
    try:
        filepath = str(filedialog.askopenfilename())
        print(filepath)
        parts = filepath.split("/")
        print(parts)
        namefile = str(parts[-1])
        pathfile = parts[:-1]

        pathfileexcel = str()
        print(pathfile)
        for number in range(0,len(pathfile)):
            pathfileexcel +=  pathfile[number] + "/"

        print(pathfileexcel)
        print(namefile)
        #reading excel file
        os.chdir(pathfileexcel)
        trialDSfile = csv.reader(open(namefile,"r"))
        #getting the DS in the file
        DStrialnodes = []
        for lines in trialDSfile:
            DStrialnodes.append(lines[0])

        previewtrialDS.delete(0, END)
        for DStrialnode in DStrialnodes:
            previewtrialDS.insert(END,DStrialnode)
        os.chdir(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\EU_Flex\PendingBlocks')

    except:
        messagebox.showinfo('Error', 'Make sure you are updating a .csv excel file and that DS nodes are pasted in column "A"')
    
    def saveDStrial(DStrialnodes):
        try:
            DSintrialfile = open("DS in trial list.csv","w",newline='')
            writer = csv.writer(DSintrialfile)
            for DStrialnode in DStrialnodes:
                tup = (DStrialnode, " ")
                writer.writerow(tup)
            DSintrialfile.close()
            DStrialsave.destroy()
            
            data = {"Content": "/md \n**DS On Trial List has been updated by " +user_login +"**"}
        
            sendwebhookmanagers(data)
            
            messagebox.showinfo('Success', 'DS in trial list has been saved')
            

        except:
            messagebox.showinfo('Error', 'Please ensure DS in trial list file exists and is not opened by anyone \n File location: Business Analyses\CentralOPS\EU_Flex\PendingBlocks')



    DStrialsave = Button(frammetrialDS, text= "Save", command=lambda:saveDStrial(DStrialnodes))
    DStrialsave.grid(row=7, column=0)
    






#######################################################################################
################################### Main tab ##########################################
#######################################################################################

#getting DS in trial
def getDStrial():
    global DStrialnodes
    DSintrialfile = csv.reader(open("DS in trial list.csv","r"))
    DStrialnodes = []
    for row in DSintrialfile:
        DStrialnodes.append(row[0])
    print(DStrialnodes)

    previewtrialDS.delete(0, END)
    for DStrialnode in DStrialnodes:
        previewtrialDS.insert(END,DStrialnode)





def mainframewidgets():
    #these are the lists
    global cycle1DSlistmax
    global cycle1LVDSlistmax
    global SDDSlistmax
    global RTSDSlistmax
    global ADHOCDSlistmax
    


    #list for the DS that will be in listboxs max
    cycle1DSlistmax = []
    cycle1LVDSlistmax = []
    SDDSlistmax = []
    RTSDSlistmax = []
    ADHOCDSlistmax = []

    #these are the listboxes
    global listboxcycle1
    global listboxcycle1max
    global listboxcycle1LV
    global listboxcycle1LVmax
    global listboxSD
    global listboxSDmax
    global listboxADHOC
    global listboxADHOCmax
    global listboxRTS
    global listboxRTSmax

    global sendmessagetochimeroom

    
    



    #Creating frame for C1
    spacerow0 = Label(framemain, text =" ").grid(row=0,column=2)
    framecycle1 = Frame(framemain)
    framecycle1.grid(row=1, column=0, columnspan=2)
    textframecycle1 = Label(framecycle1, text ="Cycle 1 / Cycle 2 (" + C1servicetype+")",justify=CENTER, font = "Helvetica 10 bold").grid(row=1,column=0, columnspan=2)
    textframecycle1subtitle1 = Label(framecycle1, text = "Surged",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=0, columnspan=1)
    textframecycle1subtitle2 = Label(framecycle1, text = "At max",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=1, columnspan=1)
    listboxcycle1 = Listbox(framecycle1)
    listboxcycle1.grid(row=3, column=0,columnspan=1)
    listboxcycle1max = Listbox(framecycle1)
    listboxcycle1max.grid(row=3, column=1, columnspan=1)
    spacecolumn3 = Label(framecycle1, text =" ").grid(row=1,column=3)

    #Creating frame for C1 AFLV
    framecycle1LV = Frame(framemain)
    framecycle1LV.grid(row=1, column=3, columnspan=2)
    textframecycle1LV = Label(framecycle1LV, text ="Cycle 1 / Cycle 2 (" + C1bulkservicetype+")",justify=CENTER, font = "Helvetica 10 bold").grid(row=1,column=3, columnspan=2)
    textframecycle1LVsubtitle1 = Label(framecycle1LV, text = "Surged",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=3, columnspan=1)
    textframecycle1LVsubtitle2 = Label(framecycle1LV, text = "At max",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=4, columnspan=1)
    listboxcycle1LV = Listbox(framecycle1LV)
    listboxcycle1LV.grid(row=3, column=3,columnspan=1)
    listboxcycle1LVmax = Listbox(framecycle1LV)
    listboxcycle1LVmax.grid(row=3, column=4, columnspan=1)

    spacecolumn3 = Label(framecycle1LV, text =" ").grid(row=1,column=5)

    #Creating frame for SD
    frameSD = Frame(framemain)
    frameSD.grid(row=1, column=6,columnspan=2)
    textframeSD = Label(frameSD, text = "SD_B / SD_C ("+ SDservicetype +")",justify=CENTER, font = "Helvetica 10 bold").grid(row=1,column=6,columnspan=2)
    textframeSDsubtitle1 = Label(frameSD, text = "Surged",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=6, columnspan=1)
    textframeSDsubtitle2 = Label(frameSD, text = "At max",justify=CENTER, font = "Helvetica 8 bold").grid(row=2,column=7, columnspan=1)
    listboxSD = Listbox(frameSD)
    listboxSD.grid(row=3, column=6,columnspan=1)
    listboxSDmax = Listbox(frameSD)
    listboxSDmax.grid(row=3, column=7,columnspan=1)

    #Creating frame for Adhoc
    spacerow6 = Label(framemain, text =" ").grid(row=6,column=2)
    frameADHOC = Frame(framemain)
    frameADHOC.grid(row=7, column=0,columnspan=2)
    textframeADHOC = Label(frameADHOC, text = "ADHOC1 / ADHOC2 ("+ Adhocservicetype +")",justify=CENTER, font = "Helvetica 10 bold").grid(row=7,column=0,columnspan=2, sticky=S)
    textframeADHOCsubtitle1 = Label(frameADHOC, text = "Surged",justify=CENTER, font = "Helvetica 8 bold").grid(row=8,column=0, columnspan=1)
    textframeADHOCsubtitle2 = Label(frameADHOC, text = "At max",justify=CENTER, font = "Helvetica 8 bold").grid(row=8,column=1, columnspan=1)
    listboxADHOC = Listbox(frameADHOC)
    listboxADHOC.grid(row=9, column=0,columnspan=1,pady= 5)
    listboxADHOCmax = Listbox(frameADHOC)
    listboxADHOCmax.grid(row=9, column=1,columnspan=1,pady= 5)

    #Creating frame for RTS
    frameRTS = Frame(framemain)
    frameRTS.grid(row=7, column=3,columnspan=2)
    textframeRTS = Label(frameRTS, text = "  RTS_1 / RTS_2 ("+ RTSservicetype+")  ",justify=CENTER, font = "Helvetica 10 bold").grid(row=7,column=3,columnspan=2, sticky=S)
    textframeRTSsubtitle1 = Label(frameRTS, text = "Surged",justify=CENTER, font = "Helvetica 8 bold").grid(row=8,column=3, columnspan=1)
    textframeRTSsubtitle2 = Label(frameRTS, text = "At max",justify=CENTER, font = "Helvetica 8 bold").grid(row=8,column=4, columnspan=1)
    listboxRTS = Listbox(frameRTS)
    listboxRTS.grid(row=9, column=3,columnspan=1,pady= 5)
    listboxRTSmax = Listbox(frameRTS)
    listboxRTSmax.grid(row=9, column=4,columnspan=1, pady= 5)
    spacerow10 = Label(framemain, text =" ").grid(row=10,column=2)

    #binding all list boxes
    listboxcycle1.bind("<1>", getlistbox)
    listboxcycle1LV.bind("<1>", getlistbox)
    listboxSD.bind("<1>", getlistbox)
    listboxADHOC.bind("<1>", getlistbox)
    listboxRTS.bind("<1>", getlistbox)
    listboxcycle1max.bind("<1>", getlistbox)
    listboxcycle1LVmax.bind("<1>", getlistbox)
    listboxSDmax.bind("<1>", getlistbox)
    listboxADHOCmax.bind("<1>", getlistbox)
    listboxRTSmax.bind("<1>", getlistbox)

    #adding other buttons
    sendmessagetochimeroom = Button(framemain, text = "Send",state=DISABLED ,command=chimemessage, width= 30)
    sendmessagetochimeroom.grid(row=7, column=6, columnspan=2, sticky=S, pady=5)

    atmaxbutton = Button(framemain, text = "At max", fg = "green", command=atmax, width= 15)
    atmaxbutton.grid(row=7,column=6)

    removebutton = Button(framemain, text = "Remove", fg = "red", command=removefrommax, width= 15)
    removebutton.grid(row=7,column=7)

    surgepriceinfobutton = Button(framemain, text = "Surge Price Info" ,command=surgepriceinfo, width= 15)
    surgepriceinfobutton.grid(row=7, column=6, columnspan=1, sticky=N)

    refreshbutton = Button(framemain, text = "Refresh" ,command=refresh, width= 15)
    refreshbutton.grid(row=7, column=7, columnspan=1, sticky=N)


# Button for getting Surge Price Information
def surgepriceinfo():
    surgeprice = csv.reader(open("Surge Price Information.csv","r"))
    for row in surgeprice:
        if "C1" == row[0]:
            info = row
            infopriceC1 = info[1]
    surgeprice = csv.reader(open("Surge Price Information.csv","r"))
    for row in surgeprice:
        if "RTS ADHOC" == row[0]:
            info = row
            infopriceRTSADHOC = info[1]
    surgeprice = csv.reader(open("Surge Price Information.csv","r"))
    for row in surgeprice:
        if "SD" == row[0]:
            info = row
            infopriceSD = info[1]
    surgeprice = csv.reader(open("Surge Price Information.csv","r"))
    for row in surgeprice:
        if "TRIAL" == row[0]:
            info = row
            infopriceTRIAL = info[1]
    
    surgewindow = Toplevel()
    titletop = Label(surgewindow, text = "Surge Price Informaton", justify=CENTER, font = "Helvetica 11 bold").grid(row=0,column=0,padx=10, columnspan=2,pady=15)
    c1surgepricetitle = Label(surgewindow, text = "Cycle 1", justify=CENTER, font = "Helvetica 9 bold").grid(row=1,column=0,padx=10, columnspan=2)
    infopriceC1label = Label(surgewindow, justify=CENTER,text = infopriceC1)
    infopriceC1label.grid(row=2,column=0, pady=5,padx=5, columnspan=2)

    RTSADHOCsurgepricetitle = Label(surgewindow, text = "RTS / ADHOC", justify=CENTER, font = "Helvetica 9 bold").grid(row=3,column=0,padx=10, columnspan=2)
    infopriceRTSADHOClabel = Label(surgewindow, justify=CENTER,text = infopriceRTSADHOC)
    infopriceRTSADHOClabel.grid(row=4,column=0, pady=5,padx=5, columnspan=2)

    SDsurgepricetitle = Label(surgewindow, text = "SD", justify=CENTER, font = "Helvetica 9 bold").grid(row=5,column=0,padx=10, columnspan=2)
    infopriceSDlabel = Label(surgewindow, justify=CENTER,text = infopriceSD)
    infopriceSDlabel.grid(row=6,column=0, pady=5,padx=5, columnspan=2)

    TRIALsurgepricetitle = Label(surgewindow, text = "TRIAL DS", justify=CENTER, font = "Helvetica 9 bold").grid(row=7,column=0,padx=10, columnspan=2)
    infopriceTRIALlabel = Label(surgewindow, justify=CENTER,text = infopriceTRIAL)
    infopriceTRIALlabel.grid(row=8,column=0, pady=5,padx=5, columnspan=2)
    
    
    
    def surgeedit():
        surgewindowedit = Toplevel()
        c1surgepricearealabel = Label(surgewindowedit,text = "Cycle 1",font = "Helvetica 9 bold").grid(row=1,column=0,padx=10,pady=5)
        c1surgepricearea = scrolledtext.ScrolledText(surgewindowedit, wrap = WORD, width = 40, height = 4, font = ("Helvetica", 8))
        c1surgepricearea.grid(row=1,column=1)
        c1surgepricearea.insert(INSERT,infopriceC1)
        
        RTSADHOCsurgepricearealabel = Label(surgewindowedit,text = "RTS / ADHOC",font = "Helvetica 9 bold").grid(row=2,column=0,padx=10,pady=5)
        RTSADHOCsurgepricearea = scrolledtext.ScrolledText(surgewindowedit, wrap = WORD, width = 40, height = 4, font = ("Helvetica", 8))
        RTSADHOCsurgepricearea.grid(row=2,column=1)
        RTSADHOCsurgepricearea.insert(INSERT,infopriceRTSADHOC)

        SDsurgepricearealabel = Label(surgewindowedit,text = "SD",font = "Helvetica 9 bold").grid(row=3,column=0,padx=10,pady=5)
        SDsurgepricearea = scrolledtext.ScrolledText(surgewindowedit, wrap = WORD, width = 40, height = 4, font = ("Helvetica", 8))
        SDsurgepricearea.grid(row=3,column=1)
        SDsurgepricearea.insert(INSERT,infopriceSD)

        TRIALsurgepricearealabel = Label(surgewindowedit,text = "TRIAL DS",font = "Helvetica 9 bold").grid(row=4,column=0,padx=10,pady=5)
        TRIALsurgepricearea = scrolledtext.ScrolledText(surgewindowedit, wrap = WORD, width = 40, height = 4, font = ("Helvetica", 8))
        TRIALsurgepricearea.grid(row=4,column=1)
        TRIALsurgepricearea.insert(INSERT,infopriceTRIAL)

        def updateinfosurge():
            try:
                surgeinfofile = open("Surge Price Information.csv","w",newline='')
                writer = csv.writer(surgeinfofile)
                
                c1surgeprice = c1surgepricearea.get('1.0', END)
                RTSADHOCsurgeprice = RTSADHOCsurgepricearea.get('1.0', END)
                SDsurgeprice= SDsurgepricearea.get('1.0', END)
                TRIALsurgeprice = TRIALsurgepricearea.get('1.0', END)
                
                tupc1 = ("C1", c1surgeprice)
                tupRTSADHOC = ("RTS ADHOC",RTSADHOCsurgeprice)
                tupSD = ("SD", SDsurgeprice)
                tupTRIAL = ("TRIAL",TRIALsurgeprice)


                writer.writerow(tupc1)
                writer.writerow(tupRTSADHOC)
                writer.writerow(tupSD)
                writer.writerow(tupTRIAL)
                
                surgeinfofile.close()
                
                messagebox.showinfo('Success', 'DS in trial list has been saved')
            

            except:
                messagebox.showinfo('Error', 'Please ensure Surge Price Information file exists and is not opened by anyone \n File location: Business Analyses\CentralOPS\EU_Flex\PendingBlocks')

        buttonupdatesurge = Button(surgewindowedit, text = "Update Info", command=updateinfosurge)
        buttonupdatesurge.grid(row=5,column=1,padx=5, pady=15,columnspan=2)
    
    buttoneditsurge = Button(surgewindow, text = "Modify Info", command=surgeedit)
    buttoneditsurge.grid(row=9,column=0,padx=5, pady=15, columnspan=2)
    
    
    
    


def chimemessage():
    sendmessagetochimeroom.config(text= "Send", state=DISABLED)
    
    #the lists "max" are the only ones who do not reset to [] every 30 mins. Therefore some of them have empty values with " " that must be removed before sending the chime message
    def removespacesinlist(listtoremovespaces):
        try:
            for numberofloops in range(1,30):
                listtoremovespaces.remove(" ")
        except:
            pass
    
    removespacesinlist(cycle1DSlistmax), removespacesinlist(cycle1LVDSlistmax), removespacesinlist(SDDSlistmax), removespacesinlist(SDDSlistmax), removespacesinlist(ADHOCDSlistmax), removespacesinlist(RTSDSlistmax)
    
    #to create the tables correctly in chime we need to get the longest list and we need to fill with " " the other lists to make them the same length
    maxlength = max(len(cycle1DSlistmax),len(cycle1LVDSlistmax),len(SDDSlistmax),len(ADHOCDSlistmax),len(RTSDSlistmax),len(cycle1DSlist),len(cycle1LVDSlist),len(SDDSlist),len(ADHOCDSlist),len(RTSDSlist),len(cycle1DSlisttrial),len(cycle1LVDSlisttrial),len(SDDSlisttrial),len(ADHOCDSlisttrial),len(RTSDSlisttrial))
    
    print(maxlength)

    def ziplist(listtoaddspaces):
        while len(listtoaddspaces)<maxlength:
            listtoaddspaces.append(" ")
    
    ziplist(cycle1DSlistmax), ziplist(cycle1LVDSlistmax), ziplist(SDDSlistmax), ziplist(SDDSlistmax), ziplist(ADHOCDSlistmax), ziplist(RTSDSlistmax), ziplist(cycle1DSlist), ziplist(cycle1LVDSlist), ziplist(SDDSlist), ziplist(ADHOCDSlist), ziplist(RTSDSlist), ziplist(cycle1DSlisttrial), ziplist(cycle1LVDSlisttrial), ziplist(SDDSlisttrial), ziplist(ADHOCDSlisttrial), ziplist(RTSDSlisttrial)

    
    

    
    urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/ed7e5ed2-d5b6-46dd-a192-3bc82a5308f0?token=NGF3UjNaQkx8MXxEVGVwLXI3U1JzU0V2bVdMNTd2Q3FaT3JYZUlqYUNvdWhuOExvSTZfUW5J"
    
    #formatting message
    datatable = ""
    for c1DS, c1LVDS, SDDS, ADHOCDS, RTSDS,c1DSmax, c1LVDSmax, SDDSmax, ADHOCDSmax, RTSDSmax, c1DStrial, c1LVDStrial, SDDStrial, ADHOCDStrial, RTSDStrial  in zip(cycle1DSlist,cycle1LVDSlist, SDDSlist, ADHOCDSlist, RTSDSlist,cycle1DSlistmax,cycle1LVDSlistmax, SDDSlistmax, ADHOCDSlistmax, RTSDSlistmax, cycle1DSlisttrial,cycle1LVDSlisttrial, SDDSlisttrial, ADHOCDSlisttrial, RTSDSlisttrial):
        datatable += "\n|" + str(c1DS) + "|" + str(c1DStrial) + "|"+ str(c1DSmax) + "|  |"+ str(c1LVDS) + "|"+ str(c1LVDStrial) + "|"+ str(c1LVDSmax)+ "|  |"+ str(SDDS) + "|"+ str(SDDStrial) + "|"+ str(SDDSmax)+ "|  |" + str(ADHOCDS) + "|"+ str(ADHOCDStrial) + "|"+ str(ADHOCDSmax)+ "|  |" + str(RTSDS)+ "|" + str(RTSDStrial)+ "|" + str(RTSDSmax) + "|"
        
    
    data = {"Content": "/md \n**Every 30 minutes Surge Monitor: **\n\n|C1 (AF Car+)| | | |C1 (AF LV)| | | |SD| | | |ADHOC| | | |RTS| |"+ timefileupdated +"|\n|-----------|-------------|------------|---|-----------|-------------|------------|---|-----------|-------------|------------|---|-----------|-------------|------------|---|-----------|-------------|------------|\n|*Surged*|*Auto-Trial*|*At Max*| |*Surged*|*Auto-Trial*|*At Max*| |*Surged*|*Auto-Trial*|*At Max*| |*Surged*|*Auto-Trial*|*At Max*| |*Surged*|*Auto-Trial*|*At Max*|"+ datatable }
    
    
    
    
    result = False
    try:
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        print("\nFailed to send Chime message: ", e)
        return result




#to send webhooks for DS that are in trial
def sendwebhooktrial(DScode,starttime,fillrate,cycle):
    data = {"Content": "/md \n**DS On Trial Warning**\n" + str(DScode) +" has pending blocks in " + str(cycle) + " at "+ str(starttime) + "\nFill rate is " + str(fillrate*100)[0:4] + "%\n**"+ str(DScode) + " has now been surged**"}
    urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/ed7e5ed2-d5b6-46dd-a192-3bc82a5308f0?token=NGF3UjNaQkx8MXxEVGVwLXI3U1JzU0V2bVdMNTd2Q3FaT3JYZUlqYUNvdWhuOExvSTZfUW5J"
    result = False
    try:
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        print("\nFailed to send Chime message: ", e)
        return result

#to send webhooks informing that in this round there is no DS that needs to be surged
def sendwebhooknoDStosurge(data):
    urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/ed7e5ed2-d5b6-46dd-a192-3bc82a5308f0?token=NGF3UjNaQkx8MXxEVGVwLXI3U1JzU0V2bVdMNTd2Q3FaT3JYZUlqYUNvdWhuOExvSTZfUW5J"
    result = False
    try:
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        print("\nFailed to send Chime message: ", e)
        return result

#when updating parametres or DS in trial
def sendwebhookmanagers(data):
    urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/54447ab4-0f49-48aa-9d9f-e46552f7d367?token=UGdHME1FQkh8MXxrU2R3X3RqU0I4aDRkWm1YdFo5ZWw5M3BSNWhpZ1FlOWExMTA3NG8xdTdB"
    result = False
    try:
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        print("\nFailed to send Chime message: ", e)
        return result






#refreshing the data, to get the latest values from flex Fill_Rate file
def getDSupdated():
    global timefileupdated

    #list for the DS that will be in listboxs max
    global cycle1DSlist
    global cycle1LVDSlist
    global SDDSlist
    global RTSDSlist
    global ADHOCDSlist
    global cycle1DSlisttrial
    global cycle1LVDSlisttrial
    global SDDSlisttrial
    global RTSDSlisttrial
    global ADHOCDSlisttrial
    
    cycle1DSlist = []
    cycle1LVDSlist = []
    SDDSlist = []
    RTSDSlist = []
    ADHOCDSlist = []
    #for trial
    cycle1DSlisttrial = []
    cycle1LVDSlisttrial = []
    SDDSlisttrial = []
    RTSDSlisttrial = []
    ADHOCDSlisttrial = []

    #getting time excel file was updated by flex team
    surgeupdated = 'Flex_Fill_Data.xlsx'
    wb = openpyxl.load_workbook(surgeupdated, data_only=True) #load workbook and telling Python to read only
    wsheet = wb.worksheets[0]
    try:
        timefileupdated = datetime.strftime(wsheet["A2"].value,"%H:%M")
    except:
        pass
    
    
    #getting DS with fill rate
    try:
        cycle1DS = []
        cycle1starttime = []
        cycle1fillrate = []
        cycle1servicetype = []
        rangeND = wsheet.iter_rows()
        for row in rangeND:
            for cell in row:
                if (cell.value == 'CYCLE_1'):
                    if wsheet.cell(row=cell.row, column=5).value == "AmFlex Car+":
                        cycle1DS.append(wsheet.cell(row=cell.row, column=3).value)
                        cycle1starttime.append((wsheet.cell(row=cell.row, column=10).value).strftime("%H:%M"))
                        cycle1fillrate.append(wsheet.cell(row=cell.row, column=9).value)
                        cycle1servicetype.append(wsheet.cell(row=cell.row, column=5).value)


        cycle1LVDS = []
        cycle1LVstarttime = []
        cycle1LVfillrate = []
        cycle1LVservicetype = []

        rangeND = wsheet.iter_rows()
        for row in rangeND:
            for cell in row:
                if (cell.value == 'CYCLE_1'):
                    if wsheet.cell(row=cell.row, column=5).value == "AmFlex Large Vehicle":
                        cycle1LVDS.append(wsheet.cell(row=cell.row, column=3).value)
                        cycle1LVstarttime.append((wsheet.cell(row=cell.row, column=10).value).strftime("%H:%M"))
                        cycle1LVfillrate.append(wsheet.cell(row=cell.row, column=9).value)
                        cycle1LVservicetype.append(wsheet.cell(row=cell.row, column=5).value)


        cycleSDDS = []
        cycleSDstarttime = []
        cycleSDfillrate = []
        cycleSDservicetype = []

        rangeSD = wsheet.iter_rows()
        for row in rangeSD:
            for cell in row:
                if (cell.value == 'CYCLE_SD_A' or cell.value == 'CYCLE_SD_B' or cell.value == 'CYCLE_SD_C'):
                    cycleSDDS.append(wsheet.cell(row=cell.row, column=3).value)
                    cycleSDstarttime.append((wsheet.cell(row=cell.row, column=10).value).strftime("%H:%M"))
                    cycleSDfillrate.append(wsheet.cell(row=cell.row, column=9).value)
                    cycleSDservicetype.append(wsheet.cell(row=cell.row, column=5).value)

        cycleADHOCDS = []
        cycleADHOCstarttime = []
        cycleADHOCfillrate = []
        cycleADHOCservicetype = []

        rangeADHOC = wsheet.iter_rows()
        for row in rangeADHOC:
            for cell in row:
                if (cell.value == 'AD_HOC_1' or cell.value == 'AD_HOC_2'):
                    cycleADHOCDS.append(wsheet.cell(row=cell.row, column=3).value)
                    cycleADHOCstarttime.append((wsheet.cell(row=cell.row, column=10).value).strftime("%H:%M"))
                    cycleADHOCfillrate.append(wsheet.cell(row=cell.row, column=9).value)
                    cycleADHOCservicetype.append(wsheet.cell(row=cell.row, column=5).value)

        cycleRTSDS = []
        cycleRTSstarttime = []
        cycleRTSfillrate = []
        cycleRTSservicetype = []

        rangeRTS = wsheet.iter_rows()
        for row in rangeRTS:
            for cell in row:
                if (cell.value == 'RTS_1' or cell.value == 'RTS_2'):
                    cycleRTSDS.append(wsheet.cell(row=cell.row, column=3).value)
                    cycleRTSstarttime.append((wsheet.cell(row=cell.row, column=10).value).strftime("%H:%M"))
                    cycleRTSfillrate.append(wsheet.cell(row=cell.row, column=9).value)
                    cycleRTSservicetype.append(wsheet.cell(row=cell.row, column=5).value)
    except:
        messagebox.showinfo('Error', 'Please ensure file Flex_Fill_Data exists \n File location: Business Analyses\CentralOPS\EU_Flex\PendingBlocks')




    #cleaning up listboxes from previous runs
    listboxcycle1.delete(0, END)
    listboxcycle1LV.delete(0, END)
    listboxSD.delete(0, END)
    listboxADHOC.delete(0, END)
    listboxRTS.delete(0, END)




    ######preparing the first filter####
    #when to surge, manual surge DS
    timesurgeNOTSD = datetime.now() + timedelta(minutes=int(NDstarttimesurge))
    timesurgeSD = datetime.now() + timedelta(minutes=int(SDstarttimesurge))
    timesurgeNOTSD = timesurgeNOTSD.strftime("%H:%M")
    timesurgeSD = timesurgeSD.strftime("%H:%M") 
    #when to surge trial
    timesurgeNOTSDtrial = datetime.now() + timedelta(minutes=int(NDstarttimesurgetrial))
    timesurgeSDtrial = datetime.now() + timedelta(minutes=int(SDstarttimesurgetrial))
    timesurgeNOTSDtrial = timesurgeNOTSDtrial.strftime("%H:%M")
    timesurgeSDtrial = timesurgeSDtrial.strftime("%H:%M") 


    ######## applying filters ###########
    #For DS that require Manual Surge
    #attaching DS to tables for each cycle
    #cycle1
    for DScode, starttime,fillrate,servicetype in zip(cycle1DS,cycle1starttime,cycle1fillrate,cycle1servicetype):
        if DScode in DStrialnodes and timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in cycle1DSlisttrial:
            cycle1DSlisttrial.append(DScode)
        if timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in DStrialnodes and DScode not in cycle1DSlistmax and DScode not in cycle1DSlist: #we dont want DS to get pulled if they are already on max from previous runs
            # print(DScode,starttime,fillrate)
            listboxcycle1.insert(END,DScode)
            cycle1DSlist.append(DScode)

    #cycle1LV
    for DScode, starttime,fillrate,servicetype in zip(cycle1LVDS,cycle1LVstarttime,cycle1LVfillrate,cycle1LVservicetype):
        if DScode in DStrialnodes and timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in cycle1LVDSlisttrial:
            cycle1LVDSlisttrial.append(DScode)
        if timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in DStrialnodes and DScode not in cycle1LVDSlistmax and DScode not in cycle1LVDSlist:
            # print(DScode,starttime,fillrate)
            listboxcycle1LV.insert(END,DScode)
            cycle1LVDSlist.append(DScode)

    #cycleSD
    for DScode, starttime,fillrate,servicetype in zip(cycleSDDS,cycleSDstarttime,cycleSDfillrate,cycleSDservicetype):
        if DScode in DStrialnodes and timesurgeSD > starttime and int(100*fillrate) < int(SDfillrate) and DScode not in SDDSlisttrial:
            SDDSlisttrial.append(DScode)
        if timesurgeSD > starttime and int(100*fillrate) < int(SDfillrate) and DScode not in DStrialnodes and DScode not in SDDSlistmax and DScode not in SDDSlist:
            # print(DScode,starttime,fillrate)
            listboxSD.insert(END,DScode)
            SDDSlist.append(DScode)

    #cycleAdhoc
    for DScode, starttime,fillrate,servicetype in zip(cycleADHOCDS,cycleADHOCstarttime,cycleADHOCfillrate,cycleADHOCservicetype):
        if DScode in DStrialnodes and timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in ADHOCDSlisttrial:
            ADHOCDSlisttrial.append(DScode)
        if timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in DStrialnodes and DScode not in ADHOCDSlistmax and DScode not in ADHOCDSlist:
            # print(DScode,starttime,fillrate)
            listboxADHOC.insert(END,DScode)
            ADHOCDSlist.append(DScode)

    #cycleRTS
    for DScode, starttime,fillrate,servicetype in zip(cycleRTSDS,cycleRTSstarttime,cycleRTSfillrate,cycleRTSservicetype):
        if DScode in DStrialnodes and timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in RTSDSlisttrial:
            RTSDSlisttrial.append(DScode)
        if timesurgeNOTSD > starttime and int(100*fillrate) < int(NDfillrate) and DScode not in DStrialnodes and DScode not in RTSDSlistmax and DScode not in RTSDSlist:
            # print(DScode,starttime,fillrate)
            listboxRTS.insert(END,DScode)
            RTSDSlist.append(DScode)
    

    #### For DS that are on Trial ####
    def addDStrialatmax(DScode,cycle):
        if cycle == "Cycle 1 AmFlex Car +":
            listboxcycle1max.insert(END,DScode)
            #updating lists
            appendinglist(cycle1DSlistmax,DScode)
        if cycle == "Cycle 1 AmFlexLarge Van":
            listboxcycle1LVmax.insert(END,DScode)
            #updating lists
            appendinglist(cycle1LVDSlistmax,DScode)
        if cycle == "SD":
            listboxSDmax.insert(END,DScode)
            #updating lists
            appendinglist(SDDSlistmax,DScode)
        if cycle == "ADHOC":
            listboxADHOCmax.insert(END,DScode)
            #updating lists
            appendinglist(ADHOCDSlistmax,DScode)
        if cycle == "RTS":
            listboxRTSmax.insert(END,DScode)
            #updating lists
            appendinglist(RTSDSlistmax,DScode)
        
    
    
    #the scheduler will receive this only for DS that are in trial and might need to be surged
    def warningtrial(DScode,starttime,fillrate,cycle):
        # if counter <2:
        answer = messagebox.askyesno("DS on trial Warning",str(DScode) +" has pending blocks in " + str(cycle) + " at "+ str(starttime) + "\nFill rate is " + str(fillrate*100)[0:4] + "%\nThis message is now copied in your clipboard. \nPlease paste message in UK Am Flex Chime Room before increasing surge \nIf the situation is really bad, do increase surge\n\nCan " + str(DScode) + " be manually surged?")
        pyperclip.copy("Hello team, " + str(DScode) + " is on trial and it has poor acceptance on " + str(cycle) + " with a fill rate of "+ str(fillrate*100)[0:4] +"% starting at "+ str(starttime))
        if answer:
            answer = messagebox.askyesno("Surging DS","Is "+ str(DScode) +" now at surge max?")
            print(answer)
            if answer:
                addDStrialatmax(DScode,cycle)
                sendwebhooktrial(DScode,starttime,fillrate,cycle)
            else:
                sendwebhooktrial(DScode,starttime,fillrate,cycle)


            
    
    #####applying filters for DS that are on trial ###########
    #cycle1
    for DScode, starttime,fillrate,servicetype in zip(cycle1DS,cycle1starttime,cycle1fillrate,cycle1servicetype):
        if timesurgeNOTSDtrial > starttime and int(100*fillrate) < int(NDfillratetrial) and DScode in DStrialnodes and DScode not in cycle1DSlistmax:
            warningtrial(DScode,starttime,fillrate,str("Cycle 1 AmFlex Car +"))
            
            
    #cycle1LV
    for DScode, starttime,fillrate,servicetype in zip(cycle1LVDS,cycle1LVstarttime,cycle1LVfillrate,cycle1LVservicetype):
        if timesurgeNOTSDtrial > starttime and int(100*fillrate) < int(NDfillratetrial) and DScode in DStrialnodes and DScode not in cycle1LVDSlistmax:
            warningtrial(DScode,starttime,fillrate, str("Cycle 1 AmFlexLarge Van"))

    #cycleSD
    for DScode, starttime,fillrate,servicetype in zip(cycleSDDS,cycleSDstarttime,cycleSDfillrate,cycleSDservicetype):
        if timesurgeSDtrial > starttime and int(100*fillrate) < int(SDfillratetrial) and DScode in DStrialnodes and DScode not in SDDSlistmax:
            warningtrial(DScode,starttime,fillrate,str("SD"))

    #cycleAdhoc
    for DScode, starttime,fillrate,servicetype in zip(cycleADHOCDS,cycleADHOCstarttime,cycleADHOCfillrate,cycleADHOCservicetype):
        if timesurgeNOTSDtrial > starttime and int(100*fillrate) < int(NDfillratetrial) and DScode in DStrialnodes and DScode not in ADHOCDSlistmax:
            warningtrial(DScode,starttime,fillrate,str("ADHOC"))
            

    #cycleRTS
    for DScode, starttime,fillrate,servicetype in zip(cycleRTSDS,cycleRTSstarttime,cycleRTSfillrate,cycleRTSservicetype):
        if timesurgeNOTSDtrial > starttime and int(100*fillrate) < int(NDfillratetrial) and DScode in DStrialnodes and DScode not in RTSDSlistmax:
            warningtrial(DScode,starttime,fillrate,str("RTS"))
            

#this is for getting the table of the ANCHOR selected. Then when pressing max button, python will know to which "max" table to move that DS (e.g from C1 to C1max not to SDmax)
def getlistbox(Event):
    global tableselected
    widgetvalue = str(Event.widget)
    if widgetvalue == ".!notebook.!frame.!frame.!listbox":
        tableselected = "listboxcycle1"
    if widgetvalue== ".!notebook.!frame.!frame.!listbox2":
        tableselected = "listboxcycle1max"
    if widgetvalue== ".!notebook.!frame.!frame2.!listbox":
        tableselected = "listboxcycle1LV"
    if widgetvalue== ".!notebook.!frame.!frame2.!listbox2":
        tableselected = "listboxcycle1LVmax"
    if widgetvalue== ".!notebook.!frame.!frame3.!listbox":
        tableselected = "listboxSD"
    if widgetvalue== ".!notebook.!frame.!frame3.!listbox2":
        tableselected = "listboxSDmax"
    if widgetvalue== ".!notebook.!frame.!frame4.!listbox":
        tableselected = "listboxADHOC"
    if widgetvalue== ".!notebook.!frame.!frame4.!listbox2":
        tableselected = "listboxADHOCmax"
    if widgetvalue== ".!notebook.!frame.!frame5.!listbox":
        tableselected = "listboxRTS"
    if widgetvalue== ".!notebook.!frame.!frame5.!listbox2":
        tableselected = "listboxRTSmax"
    
    print(tableselected)




#we want to add the DS one following the other, if they have values in the list " " we want to override them
def appendinglist(listtoupdate, DSnode):
    if " " not in listtoupdate:
        listtoupdate.append(DSnode) 
    for i in range(len(listtoupdate)):
        if listtoupdate[i] == " ":
            listtoupdate[i] = DSnode
            break     


def atmax():
    if tableselected == "listboxcycle1":
        DScode = listboxcycle1.get(ANCHOR)
        listboxcycle1max.insert(END,listboxcycle1.get(ANCHOR))
        listboxcycle1.delete(ANCHOR)
        
        #updating lists
        cycle1DSlist.remove(DScode)
        appendinglist(cycle1DSlistmax,DScode)

    if tableselected == "listboxcycle1LV":
        DScode = listboxcycle1LV.get(ANCHOR)
        listboxcycle1LVmax.insert(END,listboxcycle1LV.get(ANCHOR))
        listboxcycle1LV.delete(ANCHOR)
        
        #updating lists
        cycle1LVDSlist.remove(DScode)
        appendinglist(cycle1LVDSlistmax,DScode)

    if tableselected == "listboxSD":
        DScode = listboxSD.get(ANCHOR)
        listboxSDmax.insert(END,listboxSD.get(ANCHOR))
        listboxSD.delete(ANCHOR)

        #updating lists
        SDDSlist.remove(DScode)
        appendinglist(SDDSlistmax,DScode)
    
    if tableselected == "listboxADHOC":
        DScode = listboxADHOC.get(ANCHOR)
        listboxADHOCmax.insert(END,listboxADHOC.get(ANCHOR))
        listboxADHOC.delete(ANCHOR)

        #updating lists
        ADHOCDSlist.remove(DScode)
        appendinglist(ADHOCDSlistmax,DScode)
        

    if tableselected == "listboxRTS":
        DScode = listboxRTS.get(ANCHOR)
        listboxRTSmax.insert(END,listboxRTS.get(ANCHOR))
        listboxRTS.delete(ANCHOR)

        #updating lists
        RTSDSlist.remove(DScode)
        appendinglist(RTSDSlistmax,DScode)


def removefrommax():
    if tableselected == "listboxcycle1max":
        DScode = listboxcycle1max.get(ANCHOR)
        listboxcycle1.insert(END,listboxcycle1max.get(ANCHOR))
        listboxcycle1max.delete(ANCHOR)
        
        #updating lists
        cycle1DSlistmax.remove(DScode)
        appendinglist(cycle1DSlist,DScode)
        
        

    if tableselected == "listboxcycle1LVmax":
        DScode = listboxcycle1LVmax.get(ANCHOR)
        listboxcycle1LV.insert(END,listboxcycle1LVmax.get(ANCHOR))
        listboxcycle1LVmax.delete(ANCHOR)
        
        #updating lists
        cycle1LVDSlistmax.remove(DScode)
        appendinglist(cycle1LVDSlist,DScode)
        

    if tableselected == "listboxSDmax":
        DScode = listboxSDmax.get(ANCHOR)
        listboxSD.insert(END,listboxSDmax.get(ANCHOR))
        listboxSDmax.delete(ANCHOR)

        #updating lists
        SDDSlistmax.remove(DScode)
        appendinglist(SDDSlist,DScode)
        
    
    if tableselected == "listboxADHOCmax":
        DScode = listboxADHOCmax.get(ANCHOR)
        listboxADHOC.insert(END,listboxADHOCmax.get(ANCHOR))
        listboxADHOCmax.delete(ANCHOR)

        #updating lists
        ADHOCDSlistmax.remove(DScode)
        appendinglist(ADHOCDSlist,DScode)


    if tableselected == "listboxRTSmax":
        DScode = listboxRTSmax.get(ANCHOR)
        listboxRTS.insert(END,listboxRTSmax.get(ANCHOR))
        listboxRTSmax.delete(ANCHOR)

        #updating lists
        RTSDSlistmax.remove(DScode)
        appendinglist(RTSDSlist,DScode)


    #eliminate DS from app
    if tableselected == "listboxcycle1":
        DScode = listboxcycle1.get(ANCHOR)
        listboxcycle1.delete(ANCHOR)
        #updating lists
        cycle1DSlist.remove(DScode)
        
        

    if tableselected == "listboxcycle1LV":
        DScode = listboxcycle1LV.get(ANCHOR)
        listboxcycle1LV.delete(ANCHOR)
        #updating lists
        cycle1LVDSlist.remove(DScode)
        

    if tableselected == "listboxSD":
        DScode = listboxSD.get(ANCHOR)
        listboxSD.delete(ANCHOR)

        #updating lists
        SDDSlist.remove(DScode)

        
    
    if tableselected == "listboxADHOC":
        DScode = listboxADHOC.get(ANCHOR)
        listboxADHOC.delete(ANCHOR)

        #updating lists
        ADHOCDSlist.remove(DScode)



    if tableselected == "listboxRTS":
        DScode = listboxRTS.get(ANCHOR)
        listboxRTS.delete(ANCHOR)

        #updating lists
        RTSDSlist.remove(DScode)
    
    
    



#coundown to ping scheduler in case he/she forgets to click surge
def countdown(countnumber):
    # change text in label        
    if sendmessagetochimeroom['state'] == DISABLED:
        return
    if sendmessagetochimeroom['fg'] == "Red":
        return
    
    sendmessagetochimeroom.config(text= "Click Send " + str(countnumber) + " secs", state=ACTIVE)
    
    if countnumber > 0:
        
        root.after(1000, countdown, countnumber-1)
    
    if countnumber == 1:
        pingscheduler()

#now pinging scheduler when countdown ends
def pingscheduler():
    sendmessagetochimeroom.config(text= "Send", state=ACTIVE, fg="Red")
    data = {"Content": "/md \n**It is time to surge @" + user_login + "**"}
    urlchimeroom = "https://hooks.chime.aws/incomingwebhooks/ed7e5ed2-d5b6-46dd-a192-3bc82a5308f0?token=NGF3UjNaQkx8MXxEVGVwLXI3U1JzU0V2bVdMNTd2Q3FaT3JYZUlqYUNvdWhuOExvSTZfUW5J"
    result = False
    try:
        result = False
        session = requests.session()
        params = {'format': 'application/json'}
        response = session.post(urlchimeroom, params=params, json=data)
        if response.status_code == 200:
            result = True

        print("\nWebhooks sent\n")
        return result
        
    except Exception as e:
        print("\nFailed to send Chime message: ", e)
        return result

#this will help us to retrieve data from Flex Fill excel every 30 mins starting at 10:13
def timemonitor():
    global timenow
    global time_list
    Timer(60.0,timemonitor).start() #every 60secs the "timemonitor" function will be executed
    
    time_list1 = [datetime.strptime("10:13","%H:%M") + timedelta(minutes=30*x) for x in range(0, 21)]
    time_list=[x.strftime('%H:%M') for x in time_list1]

    
    timenow = datetime.now().strftime('%H:%M')


    if timenow in time_list:
        sendmessagetochimeroom.config(state=ACTIVE, fg="Black")
        time.sleep(1)
        countdown(600) #the countdown will be 10 mins
        getlatestparametres()
        getDStrial()
        getDSupdated()
        
        #if there is no DS to surge, lets not bother the scheduler
        if max(len(cycle1DSlistmax),len(cycle1LVDSlistmax),len(SDDSlistmax),len(ADHOCDSlistmax),len(RTSDSlistmax),len(cycle1DSlist),len(cycle1LVDSlist),len(SDDSlist),len(ADHOCDSlist),len(RTSDSlist))==0:
            data = {"Content": "/md \n**There isnt any DS that needs to be surged at the moment**"}
            sendwebhooknoDStosurge(data)
            sendmessagetochimeroom.config(text= "Send", state=DISABLED)

        #if there is DS to surge, we have to warn the scheduler
        else:
            messagebox.showinfo("Action Required","Time to surge! Tables are updated \n Do not forget to send webhook")


#this is the button that must be pressed only in emergency cases such the app closes by mistake or the scheduler was on break and is back so we can surge asap
def refresh():
    sendmessagetochimeroom.config(state=ACTIVE, fg="Black")
    getlatestparametres()
    getDStrial()
    getDSupdated()
    
    #if there is no DS to surge, lets not bother the scheduler
    if max(len(cycle1DSlistmax),len(cycle1LVDSlistmax),len(SDDSlistmax),len(ADHOCDSlistmax),len(RTSDSlistmax),len(cycle1DSlist),len(cycle1LVDSlist),len(SDDSlist),len(ADHOCDSlist),len(RTSDSlist))==0:
        data = {"Content": "/md \n**There isnt any DS that needs to be surged at the moment**"}
        sendwebhooknoDStosurge(data)
        sendmessagetochimeroom.config(text= "Send", state=DISABLED)

    #if there is DS to surge, we have to warn the scheduler
    else:
        messagebox.showinfo("Action Required","Time to surge! Tables are updated \n Do not forget to send webhook")




if __name__ == "__main__":
    getlatestparametres() #pulling parametres from excel file
    parametrestabwidgets() #building UI for the parametres tab and filling entries with the parametres pulled
    trialDSwidgets() # trial DS tab
    getDStrial() #getting the trial DS pulled from excel file
    mainframewidgets() #building UI for the main frame where all the DS with pending blocks will be shown
    getDSupdated() #pulling data from the excel file that flex team updates every 30 mins
    timemonitor() #tells the tool how often to read data and informs scheduler if webhook has not been sent after 5 mins of data being retrieved
        
    root.mainloop()










